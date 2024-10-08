import openpyxl
import requests
import time
from concurrent.futures import ThreadPoolExecutor, as_completed

# Base URL for Jenkins
jenkins_url = "http://your-jenkins-url"
username = "your-username"
api_token = "your-api-token"
poll_interval = 10  # Time in seconds to wait between status checks
max_threads = 15  # Adjust based on your server's capability and system resources

def trigger_job(job_name, params):
    url = f"{jenkins_url}/job/{job_name}/buildWithParameters"
    try:
        response = requests.post(url, params=params, auth=(username, api_token))
        if response.status_code == 201:
            print(f"Triggered {job_name}: {response.status_code}")
            location_header = response.headers.get('Location', '')
            if location_header:
                queue_id = location_header.split('/')[-2]
                return queue_id
        else:
            print(f"Failed to trigger {job_name}: {response.status_code}")
    except requests.RequestException as e:
        print(f"Exception during triggering {job_name}: {e}")
    return None

def get_build_number_from_queue(queue_id):
    url = f"{jenkins_url}/queue/item/{queue_id}/api/json"
    while True:
        try:
            response = requests.get(url, auth=(username, api_token))
            if response.status_code == 200:
                queue_info = response.json()
                if queue_info.get('executable'):
                    build_number = queue_info['executable']['number']
                    return build_number
                elif queue_info.get('cancelled'):
                    print(f"Build with queueId {queue_id} was cancelled.")
                    return None
                print(f"Waiting for job {queue_id} to start...")
            else:
                print(f"Failed to get queue info for {queue_id}: {response.status_code}")
        except requests.RequestException as e:
            print(f"Exception during getting queue info for {queue_id}: {e}")
        time.sleep(poll_interval)

def get_build_status(job_name, build_number):
    url = f"{jenkins_url}/job/{job_name}/{build_number}/api/json"
    try:
        response = requests.get(url, auth=(username, api_token))
        if response.status_code == 200:
            build_info = response.json()
            status = build_info.get('result', 'UNKNOWN')
            return status
        else:
            print(f"Failed to get status for {job_name}: {response.status_code}")
    except requests.RequestException as e:
        print(f"Exception during getting build status for {job_name}: {e}")
    return 'UNKNOWN'

def process_row(row):
    # Unpack row values and include BranchName
    app_name, job_name, env, change_request, change_task, it_release_version, obc, cbc, branch_name = row[:9]

    # Convert OBC and CBC to string and handle boolean if necessary
    if isinstance(obc, bool):
        obc = 'true' if obc else 'false'
    elif isinstance(obc, str):
        obc = obc.strip().lower()
        obc = 'true' if obc == 'yes' else 'false'
    else:
        obc = 'false'  # Default to 'false' if not a string or boolean

    if isinstance(cbc, bool):
        cbc = 'true' if cbc else 'false'
    elif isinstance(cbc, str):
        cbc = cbc.strip().lower()
        cbc = 'true' if cbc == 'yes' else 'false'
    else:
        cbc = 'false'  # Default to 'false' if not a string or boolean

    params = {
        'Env': env,
        'ChangeRequest': change_request,
        'ChangeTask': change_task,
        'ITReleaseVersion': it_release_version,
        'OBC': obc,
        'CBC': cbc,
        'BranchName': branch_name
    }

    queue_id = trigger_job(job_name, params)
    if queue_id:
        build_number = get_build_number_from_queue(queue_id)
        if build_number:
            # Polling for build status
            while True:
                status = get_build_status(job_name, build_number)
                if status in ['SUCCESS', 'FAILURE', 'UNSTABLE', 'ABORTED']:
                    break
                print(f"Waiting for build {build_number} of {job_name} to complete...")
                time.sleep(poll_interval)

            return (app_name, job_name, env, change_request, change_task, it_release_version, obc, cbc, branch_name, status)
    return (app_name, job_name, env, change_request, change_task, it_release_version, obc, cbc, branch_name, 'UNKNOWN')

def main():
    # Load the input workbook and select the active sheet
    input_wb = openpyxl.load_workbook('jobs.xlsx')
    input_ws = input_wb.active

    # Create the output workbook and sheet
    output_wb = openpyxl.Workbook()
    output_ws = output_wb.active

    # Define the header row
    headers = ['AppName', 'Job Name', 'Env', 'ChangeRequest', 'ChangeTask', 'ITReleaseVersion', 'OBC', 'CBC', 'BranchName', 'Build Status']
    output_ws.append(headers)

    # List to hold the rows from the input sheet
    rows = list(input_ws.iter_rows(min_row=2, values_only=True))

    with ThreadPoolExecutor(max_threads) as executor:
        # Submit tasks to the thread pool
        futures = [executor.submit(process_row, row) for row in rows]
        
        # Collect results as they complete
        for future in as_completed(futures):
            try:
                result = future.result()
                output_ws.append(result)
            except Exception as e:
                print(f"Exception during processing a row: {e}")

    # Save the output workbook
    output_wb.save('jobs_status.xlsx')

if __name__ == "__main__":
    main()
